import hashlib

import requests
from openpyxl import Workbook
from urllib.parse import quote
import io
import csv
import codecs
import json
from django.http import HttpResponse
from urllib.parse import quote


def download_dataset(dataset_id, file_format):
    # Получаем данные датасета из вашего API клиента
    url = f'http://127.0.0.1:8000/api/v1/datasetlist/{dataset_id}/'
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()
    num_columns = data['num_columns']
    columns = [column['name'] for column in data['dataset_columns'][:num_columns]]
    values = [[column['value'] for column in data['dataset_columns'][i:i + num_columns]] for i in
              range(0, len(data['dataset_columns']), num_columns)]
    # Получаем формат файла из данных датасета
    filename = data.get('title')

    if file_format == 'json':
        # Создаем HttpResponse и записываем данные в формате JSON
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        response.write(json.dumps(values, ensure_ascii=False, indent=4))
        return response

    elif file_format == 'csv':
        # Создаем CSV-файл и записываем данные в него
        csvfile = io.StringIO()
        writer = csv.writer(csvfile)
        # Записываем заголовки столбцов
        writer.writerow(columns)
        # Записываем значения столбцов для каждой строки
        for row in values:
            writer.writerow(row)

        # Возвращаем файл в верной кодировке
        encoded_filename = quote(filename.encode('utf-8'))
        csvfile.seek(0)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{encoded_filename}.csv"'
        response.write(codecs.BOM_UTF8)
        response.write(csvfile.read().encode('utf-8'))
        return response
    if file_format == 'sql':
        # Создаем SQL-файл и записываем данные в него
        sql_script = f"CREATE TABLE `{filename}` (\n"

        # Записываем заголовки столбцов
        for column in columns:
            sql_script += f"    `{column}` TEXT,\n"

        sql_script = sql_script[:-2]  # Удаляем последнюю запятую и новую строку
        sql_script += ");\n"

        # Записываем значения столбцов для каждой строки
        for row in values:
            sql_script += f"INSERT INTO `{filename}` VALUES ("
            for value in row:
                encoded_value = value.encode('utf-8')
                sql_script += f"'{encoded_value.decode('utf-8')}',"
            sql_script = sql_script[:-1]  # Удалить последнюю запятую
            sql_script += ");\n"

        # Создаем HttpResponse и возвращаем файл SQL
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.sql"'
        response.write(sql_script)
        return response

        # Отправляем файл SQL в ответе
        with open(sql_filename, 'rb') as file:
            response = HttpResponse(file.read(), content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="{sql_filename}"'

        return response

    elif file_format == 'xlsx':
        # Создаем XLSX-файл и записываем данные в него
        wb = Workbook()
        ws = wb.active

        # Записываем заголовки столбцов
        for col_num, column in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column)

        # Записываем значения столбцов для каждой строки
        for row_num, row in enumerate(values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Создаем BytesIO и сохраняем файл XLSX в него
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Кодируем название файла для обработки кириллических символов
        encoded_filename = quote(filename.encode('utf-8'))

        # Создаем HttpResponse и возвращаем файл XLSX
        response = HttpResponse(content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{encoded_filename}.xlsx"'
        response.write(buffer.getvalue())
        return response

    else:
        # Обработка других форматов файлов
        return HttpResponse("Unsupported file format")
